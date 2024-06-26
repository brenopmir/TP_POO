from tkinter import *
from typing import Tuple
import customtkinter
import os
import sys
# Pega o diretorio pai do arquivo
diretorioPai = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
#fornece o caminho
sys.path.append(diretorioPai)
from objects.planilha import iniciar_planilhas
from openpyxl import load_workbook

casa_path = os.path.abspath("Casa.xlsx")
casa_comodos_path = os.path.abspath("Casa_comodos.xlsx")

if not (os.path.exists(casa_path)) and os.path.exists(os.path.dirname(casa_comodos_path)):
    iniciar_planilhas()

from Interface_grafica.HeaderFrame import Header
from Interface_grafica.ComodoFrame import ComodoFrame
from Interface_grafica.DispositivosFrame import DispositivosFrame
from Interface_grafica.BotaoComodo import BotaoComodo
from Interface_grafica.LampadasFrame import LampadasFrame
from Interface_grafica.BotaoLampada import BotaoLampada
from Interface_grafica.BotaoDispositivos import BotaoDispositivo
from Interface_grafica.ArCondicionadoFrame import ArCondicionadoFrame
from Interface_grafica.BotaoArCondicionado import BotaoArCondicionado
from Interface_grafica.JanelaFrame import JanelaFrame
from Interface_grafica.BotaoJanela import BotaoJanela
from Interface_grafica.CortinaFrame import CortinaFrame
from Interface_grafica.BotaoCortina import BotaoCortina
from Interface_grafica.BotaoAdd import BotaoAdd
from Interface_grafica.BotaoRemove import BotaoRemove
from Interface_grafica.ComodoRemoveFrame import ComodoRemoveFrame
from Interface_grafica.ConfigurarLampada import ConfigurarLampadas
from Interface_grafica.ConfigurarCortina import ConfigurarCortina
from Interface_grafica.ConfigurarJanela import ConfigurarJanela
from Interface_grafica.ConfigurarAr import ConfigurarAr
from Interface_grafica.DispostivoAdd import DispositivoAddFrame
from Interface_grafica.ComodoAddFrame import ComodoAddFrame
from objects.Casa import Casa
from objects.Comodo import Comodo
from objects.ArCondicionado import Ar_Condicionado,criar_instancia_ar_condicionado
from objects.Cortinas import Cortina,criar_instancia_cortina
from objects.Janelas import Janela,criar_instancia_janela
from objects.Lampadas import Lampadas,criar_instancia_lampada
from objects.Dispositivo import Dispositivo
from typing import Type
class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()
        self.title('Automação Residencial')
        self.geometry('390x644')
        self.configure(fg_color ="#616D7A")
        self.casa = Casa("Casa")
        self.comodos = []
        self.dispositivosFrame = {}
        self.lampadasFrame ={}
        self.lampadasBotoes = {}
        self.ArCondicionadoFrame = {}
        self.ArCondicionadoBotoes = {}
        self.janelaFrame = {}
        self.janelaBotoes = {}
        self.cortinasFrame ={}
        self.cortinasBotoes = {}
        self.botoesComodo = {}
        self.lampadasConfig = {}
        self.arConfig = {}
        self.janelaConfig = {}
        self.cortinaConfig = {}
        self.frameAtual = ""
        self.comodoFrame = False
        self.contadorLampada= {}
        self.contadorArCondicionado = {}
        self.contadorJanela = {}
        self.contadorCortina = {}
        self.botaoLampada = {}
        self.botaoAr = {}
        self.botaoCortina = {}
        self.botaoJanela = {}
        self.nomeComodos=[]
        self.nomeLampadas=[]
        self.nomeArCondicionado=[]
        self.nomeJanelas=[]
        self.nomeCortina=[]
        self.botaoAddLampada = {}
        self.botaoAddAr = {}
        self.botaoAddJanela={}
        self.botaoAddCortina = {}
        self.labelDuplicataDispositivos = customtkinter.CTkLabel(self)
        self.CarregarVetores()
        self.CriarJanelas()
        
    def CarregarVetores(self):
        #Carrega todos os vetores que são lidos das planilhas
        self.nomeComodos=[]
        self.nomeLampadas=[]
        self.nomeArCondicionado=[]
        self.nomeJanelas=[]
        self.nomeCortina=[]
        
        wb=load_workbook("Casa.xlsx")
        ws_lampadas=wb["Lampadas"]
        ws_cortinas=wb["Cortinas"]
        ws_janelas=wb["Janelas"]
        ws_ares=wb["Ares Condicionados"]
        wb_Casa_comodo=load_workbook("Casa_comodos.xlsx")
        ws=wb_Casa_comodo["Comodos"]
        
        for i,row in enumerate(ws.iter_rows(min_row=2), start=1):
            self.nomeComodos.append((f"{str(row[0].value)}",f"{str(row[1].value)}"))

        for i,row in enumerate(ws_lampadas.iter_rows(min_row=2), start=1):
            self.nomeLampadas.append((f"{str(row[0].value)}",f"{str(row[1].value)}","","",f"{str(row[4].value)}","","",f"{str(row[7].value)}"))


        for i,row in enumerate(ws_cortinas.iter_rows(min_row=2), start=1):
            self.nomeCortina.append((f"{str(row[0].value)}",f"{str(row[1].value)}","","","",f"{str(row[5].value)}","",""))


        for i,row in enumerate(ws_ares.iter_rows(min_row=2), start=1):
            if(row[2].value=="True"):
                self.nomeArCondicionado.append((f"{str(row[0].value)}",f"{str(row[1].value)}","True",f"{str(row[3].value)}",f"{str(row[4].value)}","","",""))
            else:
                self.nomeArCondicionado.append((f"{str(row[0].value)}",f"{str(row[1].value)}","False",f"{str(row[3].value)}",f"{str(row[4].value)}","","",""))

        for i,row in enumerate(ws_janelas.iter_rows(min_row=2), start=1):
            if(row[6].value=="True"):
                self.nomeJanelas.append((f"{str(row[0].value)}",f"{str(row[1].value)}","","","",f"{str(row[5].value)}","True",""))
            else:
                self.nomeJanelas.append((f"{str(row[0].value)}",f"{str(row[1].value)}","","","",f"{str(row[5].value)}","False",""))


    def CriarJanelas(self):
        #Cria todos os frames e botões que serão mostrados na interface gráfica
        
        #Criando os botões que irão na página principal Comodos
        self.comodoFrame = ComodoFrame(self)
        self.comodoFrame.pack(side= "top")
        #Labels para mostrar uma mensagem em caso de erro
        self.labelDuplicata = customtkinter.CTkLabel(self.comodoFrame,
                                               text= "Já existe um comodo com esse nome, insira outro",
                                               text_color= "red")
        
        self.labelInexistente = customtkinter.CTkLabel(self.comodoFrame,
                                               text= "Esse cômodo não existe",
                                               text_color= "red")
        BotaoAdicionarComodo = BotaoAdd(self.comodoFrame, label="Adicionar novo cômodo")
        BotaoAdicionarComodo.configure(command= lambda:self.AdicionarComodoBotao())
        BotaoAdicionarComodo.pack(side = "bottom",pady= (10,10))
        
        BotaoRemoverComodo = BotaoRemove(self.comodoFrame, label="Remover cômodo")
        BotaoRemoverComodo.configure(command= lambda:self.RemoverComodoBotao())
        BotaoRemoverComodo.pack(side = "bottom",pady= (10,10))
        
        
        #Criando a pagina de cada comodo
        for nomes,numero in self.nomeComodos:
            self.contadorLampada[nomes] = 0
            self.contadorArCondicionado[nomes]  = 0
            self.contadorJanela[nomes]  = 0
            self.contadorCortina[nomes]  = 0
            
            #Criando a pagina que exibe os tipos de dispositivos
            dispositivoFrame = DispositivosFrame(master = self, nome = nomes)
            self.dispositivosFrame[nomes] = dispositivoFrame
            dispositivoFrame.header.iconeBotao.configure(command = lambda: self.VoltarFrameComodos())
            
            #Criando a pagina das lampada e configurando os botoes que irão nela
            lampadaFrame = LampadasFrame(master=self)
            lampadaFrame.header.iconeBotao.configure(command = lambda n=nomes: self.VoltarFrameDispositivosLampadas(n))
            self.lampadasFrame[nomes] = lampadaFrame
            
            self.botaoAddLampada[nomes] = BotaoAdd(self.lampadasFrame[nomes], label="Adicionar nova lâmpada")
            self.botaoAddLampada[nomes].configure(command = lambda n=nomes: self.AdicionarLampadaBotao(n))
            self.botaoAddLampada[nomes].pack(side = "bottom",pady= (10,10))
            
            for comodo,nomeLamp,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeLampadas:
                if comodo == nomes:
                    self.lampadasConfig[nomeLamp] = ConfigurarLampadas(master = self,
                                                    casas = self.casa,
                                                    comodo = nomes,
                                                    nome =nomeLamp,
                                                    intensidade=intensidade,
                                                    cor = cor
                                                    )
                    self.contadorLampada[nomes]  += 1
                    self.lampadasConfig[nomeLamp].header.iconeBotao.configure(command = lambda n=nomes: self.VoltarFrameLampadas(n))
                    self.lampadasConfig[nomeLamp].excluir.configure(command = lambda n1=nomeLamp,n2=nomes:self.RemoverLampada(nomeLampada=n1,nomeComodo=n2))
                    self.lampadasBotoes[nomeLamp] = BotaoLampada(self.lampadasFrame[nomes],
                                                             nomeLampada=nomeLamp,
                                                             cor=cor,
                                                             intensidade=intensidade)
                    self.lampadasBotoes[nomeLamp].configure(command = lambda n=nomeLamp:self.MudarFrameConfigLampadas(n))
                    self.lampadasBotoes[nomeLamp].pack(side="top", pady= (10,10))
            
            #Criando a pagina do ar condicionado e configurando os botoes que irão nela 
            arCondicionadoFrame = ArCondicionadoFrame(master=self)
            arCondicionadoFrame.header.iconeBotao.configure(command = lambda n=nomes: self.VoltarFrameDispositivosAr(n))
            self.ArCondicionadoFrame[nomes] = arCondicionadoFrame
            
            self.botaoAddAr[nomes] = BotaoAdd(self.ArCondicionadoFrame[nomes], label="Adicionar novo ar")
            self.botaoAddAr[nomes].configure(command = lambda n=nomes: self.AdicionarArBotao(n))
            self.botaoAddAr[nomes].pack(side = "bottom",pady= (10,10))
             
            for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeArCondicionado:
                if comodo == nomes:
                    self.arConfig[nome] = ConfigurarAr(master = self,
                                            casas = self.casa,
                                            comodo = nomes,
                                            nome =nome,
                                            ligado = ligado,
                                            temperatura= temperatura,
                                            intensidade=intensidade,
                                                    )
                    self.contadorArCondicionado[nomes]  += 1
                    self.arConfig[nome].header.iconeBotao.configure(command = lambda n=nomes: self.VoltarFrameAr(n))
                    self.arConfig[nome].excluir.configure(command = lambda n1=nome,n2=nomes:self.RemoverAr(nomeAr=n1,nomeComodo=n2))
                    
                    self.ArCondicionadoBotoes[nome] = BotaoArCondicionado(self.ArCondicionadoFrame[nomes],
                                                             nomeAr=nome,
                                                             ligado=ligado,
                                                             temperatura= temperatura,
                                                             intensidade=intensidade)
                    self.ArCondicionadoBotoes[nome].configure(command = lambda n=nome:self.MudarFrameConfigAr(n))
                    self.ArCondicionadoBotoes[nome].pack(side="top", pady= (10,10))
              
              
            #Criando a pagina das janelas e configurando os botoes que irão nela        
            janelaFrame = JanelaFrame(master=self)
            janelaFrame.header.iconeBotao.configure(command = lambda n=nomes: self.VoltarFrameDispositivosJanela(n))
            self.janelaFrame[nomes] = janelaFrame
            
            self.botaoAddJanela[nomes] = BotaoAdd(self.janelaFrame[nomes], label="Adicionar nova janela")
            self.botaoAddJanela[nomes].configure(command = lambda n=nomes: self.AdicionarJanelaBotao(n))
            self.botaoAddJanela[nomes].pack(side = "bottom",pady= (10,10))
             
            for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeJanelas:
                if comodo == nomes:
                    self.janelaConfig[nome] = ConfigurarJanela(master = self,
                                                    casas= self.casa,
                                                    comodo = nomes,
                                                    nome =nome,
                                                    abertura=abertura,
                                                    trancado = tranca
                                                    )
                    self.janelaConfig[nome].header.iconeBotao.configure(command = lambda n=nomes: self.VoltarFrameJanela(n))
                    self.janelaConfig[nome].excluir.configure(command = lambda n1=nome,n2=nomes:self.RemoverJanela(nomeJanela=n1,nomeComodo=n2))
                    self.contadorJanela[nomes]  += 1
                    self.janelaBotoes[nome] = BotaoJanela(self.janelaFrame[nomes],
                                                             nomeJanela=nome,
                                                             abertura=abertura,
                                                             trancado=tranca)
                    self.janelaBotoes[nome].configure(command = lambda n=nome:self.MudarFrameConfigJanela(n))
                    self.janelaBotoes[nome].pack(side="top", pady= (10,10))
             
             
            #Criando a pagina das cortinas e configurando os botoes que irão nela 
            cortinaFrame = CortinaFrame(master=self)
            cortinaFrame.header.iconeBotao.configure(command = lambda n=nomes: self.VoltarFrameDispositivosCortina(n))
            self.cortinasFrame[nomes] = cortinaFrame
            
            self.botaoAddCortina[nomes] = BotaoAdd(self.cortinasFrame[nomes], label="Adicionar nova cortina")
            self.botaoAddCortina[nomes].configure(command = lambda n=nomes: self.AdicionarCortinaBotao(n))
            self.botaoAddCortina[nomes].pack(side = "bottom",pady= (10,10))
                    
            for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeCortina:
                if comodo == nomes:
                    self.cortinaConfig[nome] = ConfigurarCortina(master = self,
                                                    casas = self.casa,
                                                    comodo = nomes,
                                                    nome =nome,
                                                    abertura=abertura,
                                                    )
                    self.cortinaConfig[nome].excluir.configure(command = lambda n1=nome,n2=nomes:self.RemoverCortina(nomeCortina=n1,nomeComodo=n2))
                    self.cortinaConfig[nome].header.iconeBotao.configure(command = lambda n=nomes: self.VoltarFrameCortina(n))
                    self.contadorCortina[nomes]  += 1
                    self.cortinasBotoes[nome] = BotaoCortina(self.cortinasFrame[nomes],
                                                             nomeCortina=nome,
                                                             abertura=abertura)
                    self.cortinasBotoes[nome].configure(command = lambda n=nome:self.MudarFrameConfigCortina(n))
                    self.cortinasBotoes[nome].pack(side="top", pady= (10,10))        
                           
            
            #Criando os botões que irão na pagina dos dispositivos    
            self.botaoLampada[nomes] = BotaoDispositivo(self.dispositivosFrame[nomes], nomeComodo="Lâmpadas",
                                  numeroDispositivos= self.contadorLampada[nomes] ,
                                  caminho="src/icons/lampada.png")
            self.botaoLampada[nomes].configure(command = lambda n=nomes:self.MudarFrameLampadas(n))
            self.botaoLampada[nomes].pack(side="top", pady= (10,10))
            
            self.botaoAr[nomes] = BotaoDispositivo(self.dispositivosFrame[nomes], nomeComodo="Ar Condicionado",
                                  numeroDispositivos= self.contadorArCondicionado[nomes] ,
                                  caminho="src/icons/arCondicionado.png")
            self.botaoAr[nomes].configure(command = lambda n=nomes:self.MudarFrameAr(n))
            self.botaoAr[nomes].pack(side="top", pady= (10,10))
        
            self.botaoJanela[nomes] = BotaoDispositivo(self.dispositivosFrame[nomes], nomeComodo="Janela",
                                  numeroDispositivos= self.contadorJanela[nomes] ,
                                  caminho="src/icons/Janela.png")
            self.botaoJanela[nomes].configure(command = lambda n=nomes:self.MudarFrameJanela(n))
            self.botaoJanela[nomes].pack(side="top", pady= (10,10))
        
            self.botaoCortina[nomes] = BotaoDispositivo(self.dispositivosFrame[nomes], nomeComodo="Cortina",
                                  numeroDispositivos= self.contadorCortina[nomes] ,
                                  caminho="src/icons/cortina.png")
            self.botaoCortina[nomes].configure(command = lambda n=nomes:self.MudarFrameCortina(n))
            self.botaoCortina[nomes].pack(side="top", pady= (10,10))
            
            #Colocando os botões da pagina do cômodo
            self.botoesComodo[nomes] = BotaoComodo(self.comodoFrame, nomeComodo=nomes, numeroDispositivos=numero)
            self.botoesComodo[nomes].configure(command = lambda n=nomes:self.MudarFrameDispositivos(n))
            self.botoesComodo[nomes].pack(side="top", pady= (10,10))

    def RecarregarBotoesComodos(self):
        #Recarrega todos os botões presentes na página inicial cômodos, de forma a atualizar 
        #os dados dos botões de acordo com as planilhas
        for nomes,numero in self.nomeComodos:
            self.botoesComodo[nomes].destroy()
        for nomes,numero in self.nomeComodos:
            self.botoesComodo[nomes] = BotaoComodo(self.comodoFrame, nomeComodo=nomes, numeroDispositivos=numero)
            self.botoesComodo[nomes].configure(command = lambda n=nomes:self.MudarFrameDispositivos(n))
            self.botoesComodo[nomes].pack(side="top", pady= (10,10))

    def RecarregarBotoesDispositivos(self,nomeComodo):
        '''
        Recarrega todos os botões presentes na página dos dispositivos de um comodo, de forma a atualizar 
       o número de dispositivos de acordo com as planilhas
        :param nomeComodo: representa o comodo cuja pagina de dispositivos será atualizada 
        '''
        for nomes,numero in self.nomeComodos:
            self.botoesComodo[nomes].destroy()
        for nomes,numero in self.nomeComodos:
            self.botoesComodo[nomes] = BotaoComodo(self.comodoFrame, nomeComodo=nomes, numeroDispositivos=numero)
            self.botoesComodo[nomes].configure(command = lambda n=nomes:self.MudarFrameDispositivos(n))
            self.botoesComodo[nomes].pack(side="top", pady= (10,10))
        
        self.botaoLampada[nomeComodo].destroy()
        self.botaoAr[nomeComodo].destroy()
        self.botaoJanela[nomeComodo].destroy()
        self.botaoCortina[nomeComodo].destroy()
        
        
        self.botaoLampada[nomeComodo] = self.botaoLampada[nomeComodo] = BotaoDispositivo(self.dispositivosFrame[nomeComodo], nomeComodo="Lâmpadas",
                                  numeroDispositivos= self.contadorLampada[nomeComodo],
                                  caminho="src/icons/lampada.png")
        self.botaoLampada[nomeComodo].configure(command = lambda n=nomeComodo:self.MudarFrameLampadas(n))
        self.botaoLampada[nomeComodo].pack(side="top", pady= (10,10))
        
     
        self.botaoAr[nomeComodo] = self.botaoAr[nomeComodo] = BotaoDispositivo(self.dispositivosFrame[nomeComodo], nomeComodo="Ar Condicionado",
                                  numeroDispositivos= self.contadorArCondicionado[nomeComodo],
                                  caminho="src/icons/arCondicionado.png")
        self.botaoAr[nomeComodo].configure(command = lambda n=nomeComodo:self.MudarFrameAr(n))
        self.botaoAr[nomeComodo].pack(side="top", pady= (10,10))
        
        
        self.botaoJanela[nomeComodo] = self.botaoJanela[nomeComodo] = BotaoDispositivo(self.dispositivosFrame[nomeComodo], nomeComodo="Janela",
                                  numeroDispositivos= self.contadorJanela[nomeComodo],
                                  caminho="src/icons/Janela.png")
        self.botaoJanela[nomeComodo].configure(command = lambda n=nomeComodo:self.MudarFrameJanela(n))
        self.botaoJanela[nomeComodo].pack(side="top", pady= (10,10))
        
        self.botaoCortina[nomeComodo] = self.botaoCortina[nomeComodo] = BotaoDispositivo(self.dispositivosFrame[nomeComodo], nomeComodo="Cortina",
                                  numeroDispositivos= self.contadorCortina[nomeComodo],
                                  caminho="src/icons/cortina.png")
        self.botaoCortina[nomeComodo].configure(command = lambda n=nomeComodo:self.MudarFrameCortina(n))
        self.botaoCortina[nomeComodo].pack(side="top", pady= (10,10))
        
    def RecarregarBotoesLampada(self,nomeComodo):
        '''
        Recarrega todos os botões presentes na página das lampadas de um comodo, de forma a atualizar 
        os dados das lampadas de acordo com as planilhas
        :param nomeComodo: representa o comodo cuja pagina Lâmpadas será atualizada 
        '''
        self.CarregarVetores()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeLampadas:
            if comodo == nomeComodo:
                self.lampadasBotoes[nome].destroy()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeLampadas:
            if comodo == nomeComodo:
                self.lampadasBotoes[nome] = BotaoLampada(self.lampadasFrame[nomeComodo],
                                                             nomeLampada=nome,
                                                             cor=cor,
                                                             intensidade=intensidade)
                self.lampadasBotoes[nome].configure(command = lambda n=nome:self.MudarFrameConfigLampadas(n))
                self.lampadasBotoes[nome].pack(side="top", pady= (10,10))  
                
    def RecarregarBotoesAr(self,nomeComodo):
        '''
        Recarrega todos os botões presentes na página dos ares condicionados de um comodo, de forma a atualizar 
        os dados dos ares de acordo com as planilhas
        :param nomeComodo: representa o comodo cuja pagina ArCondicionado será atualizada 
        '''
        self.CarregarVetores()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeArCondicionado:
            if comodo == nomeComodo:
                self.ArCondicionadoBotoes[nome].destroy()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeArCondicionado:
            if comodo == nomeComodo:
                self.ArCondicionadoBotoes[nome] = BotaoArCondicionado(self.ArCondicionadoFrame[nomeComodo],
                                                             nomeAr=nome,
                                                             ligado=ligado,
                                                             temperatura=temperatura,
                                                             intensidade=intensidade)
                self.ArCondicionadoBotoes[nome].configure(command = lambda n=nome:self.MudarFrameConfigAr(n))
                self.ArCondicionadoBotoes[nome].pack(side="top", pady= (10,10))  
                
    def RecarregarBotoesJanela(self,nomeComodo):
        '''
        Recarrega todos os botões presentes na página das janelas de um comodo, de forma a atualizar 
        os dados das janelas de acordo com as planilhas
        :param nomeComodo: representa o comodo cuja pagina Janela será atualizada 
        '''
        self.CarregarVetores()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeJanelas:
            if comodo == nomeComodo:
                self.janelaBotoes[nome].destroy()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeJanelas:
            if comodo == nomeComodo:
                self.janelaBotoes[nome] = BotaoJanela(self.janelaFrame[nomeComodo],
                                                             nomeJanela=nome,
                                                             trancado=tranca,
                                                             abertura=abertura)
                self.janelaBotoes[nome].configure(command = lambda n=nome:self.MudarFrameConfigJanela(n))
                self.janelaBotoes[nome].pack(side="top", pady= (10,10))  
                    
    def RecarregarBotoesCortina(self,nomeComodo):
        '''
        Recarrega todos os botões presentes na página das Cortinas de um comodo, de forma a atualizar 
        os dados das Cortinas de acordo com as planilhas
        :param nomeComodo: representa o comodo cuja pagina Ares será atualizada 
        '''
        self.CarregarVetores()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeCortina:
            if comodo == nomeComodo:
                self.cortinasBotoes[nome].destroy()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeCortina:
            if comodo == nomeComodo:
                self.cortinasBotoes[nome] = BotaoCortina(self.cortinasFrame[nomeComodo],
                                                             nomeCortina=nome,
                                                             abertura= abertura)
                self.cortinasBotoes[nome].configure(command = lambda n=nome:self.MudarFrameConfigCortina(n))
                self.cortinasBotoes[nome].pack(side="top", pady= (10,10))  
    
    def AdicionarLampadaBotao(self,nomeComodo):
        '''
        Funcão para adicionar uma nova lâmpada
        :param nomeComodo: representa o comodo que tera uma lampada adicionada
        '''
        self.inputLampadaAdd = DispositivoAddFrame(self.lampadasFrame[nomeComodo])
        self.inputLampadaAdd.submit.configure(command = lambda n=nomeComodo:self.SubmeterAddLampada(n))
        self.inputLampadaAdd.pack(side="bottom")

        
        
    def SubmeterAddLampada(self,nomeComodo):
        '''
        Funcão que submete uma nova lâmpada para as planilhas e cria novos componentes para ela
        :param nomeComodo: representa o comodo que tera uma lampada adicionada
        '''
        NomeLampada = self.inputLampadaAdd.input.get()
        self.labelDuplicataDispositivos.pack_forget()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeLampadas:
            #verifica se existe um dispositivo com esse nome
            if nome == NomeLampada:
                self.labelDuplicataDispositivos = customtkinter.CTkLabel(self.lampadasFrame[nomeComodo],
                                               text= "Já existe um dispositivo com esse nome, insira outro",
                                               text_color= "red")
                self.labelDuplicataDispositivos.pack(side="top")
                return
        self.casa.comodos[nomeComodo].AdicionarDispositivo(tipo=1,nome=NomeLampada)
        self.casa.SalvarQuantidadeDeDispositivosComodo()
        self.CarregarVetores()
        
        #Cria a página de configuração para a nova lâmpada
        self.lampadasConfig[NomeLampada] = ConfigurarLampadas(master = self,
                                                    casas = self.casa,
                                                    comodo = nomeComodo,
                                                    nome =NomeLampada,
                                                    intensidade=0,
                                                    cor = "Branco"
                                                    )
        self.contadorLampada[nomeComodo]  += 1
        self.lampadasConfig[NomeLampada].header.iconeBotao.configure(command = lambda n=nomeComodo: self.VoltarFrameLampadas(n))
        self.lampadasConfig[NomeLampada].excluir.configure(command = lambda n1=NomeLampada,n2=nomeComodo:self.RemoverLampada(nomeLampada=n1,nomeComodo=n2))
                    
        self.lampadasBotoes[NomeLampada] = BotaoLampada(self.lampadasFrame[nomeComodo],
                                                             nomeLampada=NomeLampada,
                                                             cor="Branco",
                                                             intensidade=0)
        self.lampadasBotoes[NomeLampada].configure(command = lambda n=NomeLampada:self.MudarFrameConfigLampadas(n))
        self.lampadasBotoes[NomeLampada].pack(side="top", pady= (10,10))
        self.inputLampadaAdd.pack_forget()
        
        self.RecarregarBotoesDispositivos(nomeComodo)
    
    def RemoverLampada(self,nomeLampada,nomeComodo):
        '''
        Funcão para remover uma lâmpada
        :param nomeComodo: representa o comodo que tera uma lampada removida
        :param nomeLampada: representa o nome da lâmpada que será removida
        '''
        
        self.casa.comodos[nomeComodo].RemoverDispositivo(1,nomeLampada)
        self.casa.SalvarQuantidadeDeDispositivosComodo()
        self.contadorLampada[nomeComodo] -= 1
        self.lampadasBotoes[nomeLampada].destroy()
        self.lampadasConfig[nomeLampada].pack_forget()
        self.lampadasConfig[nomeLampada].destroy()
        self.lampadasFrame[nomeComodo].pack(side="top")
        self.frameAtual=nomeComodo
        self.CarregarVetores()
        self.RecarregarBotoesLampada(nomeComodo)
        self.RecarregarBotoesDispositivos(nomeComodo)
        self.RecarregarBotoesComodos()
        
    def RemoverAr(self,nomeAr,nomeComodo):
        '''
        Funcão para remover um Ar Condicionado
        :param nomeComodo: representa o comodo que terá um ar condicionado removido
        :param nomeAr: representa o nome do ar que será removida
        '''
        
        self.casa.comodos[nomeComodo].RemoverDispositivo(3,nomeAr)
        self.casa.SalvarQuantidadeDeDispositivosComodo()
        self.contadorArCondicionado[nomeComodo] -= 1
        self.ArCondicionadoBotoes[nomeAr].destroy()
        self.CarregarVetores()
        self.RecarregarBotoesAr(nomeComodo)
        self.RecarregarBotoesDispositivos(nomeComodo)
        self.RecarregarBotoesComodos() 
        self.VoltarFrameAr(nomeComodo)
        
    def AdicionarArBotao(self,nomeComodo):
        '''
        Funcão para adicionar um novo ar condicionado
        :param nomeComodo: representa o comodo que tera um ar condiconado adicionado
        '''
        self.inputArAdd = DispositivoAddFrame(self.ArCondicionadoFrame[nomeComodo])
        self.inputArAdd.submit.configure(command = lambda n=nomeComodo:self.SubmeterAddAr(n))
        self.inputArAdd.pack(side="bottom")    
        
    def SubmeterAddAr(self,nomeComodo):
        '''
        Funcão que submete um novo ar condicionado para as planilhas e cria novos componentes para ele
        :param nomeComodo: representa o comodo que tera um ar condicionado adicionada
        '''
        NomeAr = self.inputArAdd.input.get()
        self.labelDuplicataDispositivos.pack_forget()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeArCondicionado:
            #verifica se existe um dispositivo com esse nome
            if nome == NomeAr:
                self.labelDuplicataDispositivos = customtkinter.CTkLabel(self.ArCondicionadoFrame[nomeComodo],
                                               text= "Já existe um dispositivo com esse nome, insira outro",
                                               text_color= "red")
                self.labelDuplicataDispositivos.pack(side="top")
                return
        self.casa.comodos[nomeComodo].AdicionarDispositivo(tipo=3,nome=NomeAr)
        self.casa.SalvarQuantidadeDeDispositivosComodo()
        self.CarregarVetores()
        
        #Cria a página de configuração para o novo ar condicionado
        self.arConfig[NomeAr] = ConfigurarAr(master = self,
                                            casas = self.casa,
                                            comodo = nomeComodo,
                                            nome =NomeAr,
                                            ligado = "False",
                                            temperatura= 0,
                                            intensidade=0,
                                                    )
        self.contadorArCondicionado[nomeComodo]  += 1
        self.arConfig[NomeAr].header.iconeBotao.configure(command = lambda n=nomeComodo: self.VoltarFrameAr(n))
        self.arConfig[NomeAr].excluir.configure(command = lambda n1=NomeAr,n2=nomeComodo:self.RemoverAr(nomeAr=n1,nomeComodo=n2))
                    
        self.ArCondicionadoBotoes[NomeAr] = BotaoArCondicionado(self.ArCondicionadoFrame[nomeComodo],
                                                             nomeAr=NomeAr,
                                                             ligado="False",
                                                             temperatura= 20,
                                                             intensidade=0)
        self.ArCondicionadoBotoes[NomeAr].configure(command = lambda n=NomeAr:self.MudarFrameConfigAr(n))
        self.ArCondicionadoBotoes[NomeAr].pack(side="top", pady= (10,10))
        self.inputArAdd.pack_forget()
        
        self.RecarregarBotoesDispositivos(nomeComodo)
        
    def AdicionarJanelaBotao(self,nomeComodo):
        '''
        Funcão para adicionar uma nova janela
        :param nomeComodo: representa o comodo que tera uma janela adicionada
        '''
        self.inputJanelaAdd = DispositivoAddFrame(self.janelaFrame[nomeComodo])
        self.inputJanelaAdd.submit.configure(command = lambda n=nomeComodo:self.SubmeterAddJanela(n))
        self.inputJanelaAdd.pack(side="bottom")

        
        
    def SubmeterAddJanela(self,nomeComodo):
        '''
        Funcão que submete uma nova janela para as planilhas e cria novos componentes para ela
        :param nomeComodo: representa o comodo que tera uma janela adicionada
        '''
        NomeJanela = self.inputJanelaAdd.input.get()
        self.labelDuplicataDispositivos.pack_forget()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeLampadas:
            #verifica se já existe um dispositivo com esse novo
            if nome == NomeJanela:
                self.labelDuplicataDispositivos = customtkinter.CTkLabel(self.janelaFrame[nomeComodo],
                                               text= "Já existe um dispositivo com esse nome, insira outro",
                                               text_color= "red")
                self.labelDuplicataDispositivos.pack(side="top")
                return
        self.casa.comodos[nomeComodo].AdicionarDispositivo(tipo=4,nome=NomeJanela)
        self.casa.SalvarQuantidadeDeDispositivosComodo()
        self.CarregarVetores()
        
        #Cria a página de configuração para a nova janela
        self.janelaConfig[NomeJanela] = ConfigurarJanela(master = self,
                                                    casas = self.casa,
                                                    comodo = nomeComodo,
                                                    nome =NomeJanela,
                                                    abertura=0,
                                                    trancado=  "False"
                                                    )
        self.contadorJanela[nomeComodo]  += 1
        self.janelaConfig[NomeJanela].header.iconeBotao.configure(command = lambda n=nomeComodo: self.VoltarFrameJanela(n))
        self.janelaConfig[NomeJanela].excluir.configure(command = lambda n1=NomeJanela,n2=nomeComodo:self.RemoverJanela(nomeJanela=n1,nomeComodo=n2))
                    
        self.janelaBotoes[NomeJanela] = BotaoJanela(self.janelaFrame[nomeComodo],
                                                             nomeJanela=NomeJanela,
                                                             trancado="True",
                                                             abertura=0)
        self.janelaBotoes[NomeJanela].configure(command = lambda n=NomeJanela:self.MudarFrameConfigJanela(n))
        self.janelaBotoes[NomeJanela].pack(side="top", pady= (10,10))
        self.inputJanelaAdd.pack_forget()
        
        self.RecarregarBotoesDispositivos(nomeComodo)
      
    def RemoverJanela(self,nomeJanela,nomeComodo):
        '''
        Funcão para remover uma janela
        :param nomeComodo: representa o comodo que tera uma janela removida
        :param nomeJanela: representa o nome da janela que será removida
        '''
        self.casa.comodos[nomeComodo].RemoverDispositivo(4,nomeJanela)
        self.casa.SalvarQuantidadeDeDispositivosComodo()
        self.contadorJanela[nomeComodo] -= 1
        self.janelaBotoes[nomeJanela].destroy()
        self.CarregarVetores()
        self.RecarregarBotoesJanela(nomeComodo)
        self.RecarregarBotoesDispositivos(nomeComodo)
        self.RecarregarBotoesComodos() 
        self.VoltarFrameJanela(nomeComodo)  
        
    def AdicionarCortinaBotao(self,nomeComodo):
        '''
        Funcão para adicionar uma nova cortina
        :param nomeComodo: representa o comodo que tera uma cortina adicionada
        '''
        self.inputCortinaAdd = DispositivoAddFrame(self.cortinasFrame[nomeComodo])
        self.inputCortinaAdd.submit.configure(command = lambda n=nomeComodo:self.SubmeterAddCortina(n))
        self.inputCortinaAdd.pack(side="bottom")

        
        
    def SubmeterAddCortina(self,nomeComodo):
        '''
        Funcão que submete uma nova cortina para as planilhas e cria novos componentes para ela
        :param nomeComodo: representa o comodo que tera uma cortina adicionada
        '''
        NomeCortina = self.inputCortinaAdd.input.get()
        self.labelDuplicataDispositivos.pack_forget()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeCortina:
            #verifica se já existe um dispositivo com esse nome
            if nome == NomeCortina:
                self.labelDuplicataDispositivos = customtkinter.CTkLabel(self.cortinasFrame[nomeComodo],
                                               text= "Já existe um dispositivo com esse nome, insira outro",
                                               text_color= "red")
                self.labelDuplicataDispositivos.pack(side="top")
                return
        self.casa.comodos[nomeComodo].AdicionarDispositivo(tipo=2,nome=NomeCortina)
        self.casa.SalvarQuantidadeDeDispositivosComodo()
        self.CarregarVetores()
        
        #Cria a página de configuração para a nova cortina
        self.cortinaConfig[NomeCortina] = ConfigurarCortina(master = self,
                                                    casas = self.casa,
                                                    comodo = nomeComodo,
                                                    nome =NomeCortina,
                                                    abertura= 0
                                                    )
        self.contadorCortina[nomeComodo]  += 1
        self.cortinaConfig[NomeCortina].header.iconeBotao.configure(command = lambda n=nomeComodo: self.VoltarFrameCortina(n))
        self.cortinaConfig[NomeCortina].excluir.configure(command = lambda n1=NomeCortina,n2=nomeComodo:self.RemoverCortina(nomeCortina=n1,nomeComodo=n2))
                    
        self.cortinasBotoes[NomeCortina] = BotaoCortina(self.cortinasFrame[nomeComodo],
                                                             nomeCortina=NomeCortina,
                                                             abertura = 0)
        self.cortinasBotoes[NomeCortina].configure(command = lambda n=NomeCortina:self.MudarFrameConfigCortina(n))
        self.cortinasBotoes[NomeCortina].pack(side="top", pady= (10,10))
        self.inputCortinaAdd.pack_forget()
        
        self.RecarregarBotoesDispositivos(nomeComodo)
        
    def RemoverCortina(self,nomeCortina,nomeComodo):
        '''
        Funcão para remover uma cortina
        :param nomeComodo: representa o comodo que tera uma cortina removida
        :param nomeCortina: representa o nome da cortina que será removida
        '''
        self.casa.comodos[nomeComodo].RemoverDispositivo(2,nomeCortina)
        self.casa.SalvarQuantidadeDeDispositivosComodo()
        self.contadorCortina[nomeComodo] -= 1
        self.cortinasBotoes[nomeCortina].destroy()
        self.CarregarVetores()
        self.RecarregarBotoesCortina(nomeComodo)
        self.RecarregarBotoesDispositivos(nomeComodo)
        self.RecarregarBotoesComodos() 
        self.VoltarFrameCortina(nomeComodo)
        
    def AdicionarComodoBotao(self):
        #Função para adicionar um novo comodo
        self.inputComodoAdd = ComodoAddFrame(self.comodoFrame)
        self.inputComodoAdd.submit.configure(command = lambda:self.SubmeterAddComodo())
        self.inputComodoAdd.pack(side="bottom")    
        
    def SubmeterAddComodo(self):
        #funcao para enviar o novo comodo para as planilhas e coloca-lo na tela, criando
        # uma nova página de dispositivos para ele
        NomeComodo = self.inputComodoAdd.input.get()
        self.labelDuplicata.pack_forget()
        for nomes,numero in self.nomeComodos:
            if nomes == NomeComodo:
                self.labelDuplicata.pack(side = "top")
                return
        
        self.casa.AdicionarComodo(NomeComodo)
        self.CarregarVetores()
        
        
        #Criando o botao do novo comodo
        self.botoesComodo[NomeComodo] = BotaoComodo(self.comodoFrame, nomeComodo=NomeComodo, numeroDispositivos=0)
        self.botoesComodo[NomeComodo].configure(command = lambda n=NomeComodo:self.MudarFrameDispositivos(n))
        self.botoesComodo[NomeComodo].pack(side="top", pady= (10,10))

        #Criando o frame para os dispostivos do novo comodo        
        dispositivoFrame = DispositivosFrame(master = self, nome = NomeComodo)
        self.dispositivosFrame[NomeComodo] = dispositivoFrame
        dispositivoFrame.header.iconeBotao.configure(command = lambda: self.VoltarFrameComodos())
           
        #Zerando os contadores do numero de dispositivos
        self.contadorLampada[NomeComodo] = 0
        self.contadorArCondicionado[NomeComodo]  = 0
        self.contadorJanela[NomeComodo]  = 0
        self.contadorCortina[NomeComodo]  = 0 
           
        #Criando os botões que irão na pagina dos dispositivos
        #Lampadas
        self.botaoLampada[NomeComodo] = BotaoDispositivo(self.dispositivosFrame[NomeComodo], nomeComodo="Lâmpadas",
                                  numeroDispositivos= 0,
                                  caminho="src/icons/lampada.png")
        self.botaoLampada[NomeComodo].configure(command = lambda n=NomeComodo:self.MudarFrameLampadas(n))
        self.botaoLampada[NomeComodo].pack(side="top", pady= (10,10))
        lampadaFrame = LampadasFrame(master=self)
        lampadaFrame.header.iconeBotao.configure(command = lambda n=NomeComodo: self.VoltarFrameDispositivosLampadas(n))
        self.lampadasFrame[NomeComodo] = lampadaFrame
        
        self.botaoAddLampada[NomeComodo] = BotaoAdd(self.lampadasFrame[NomeComodo], label="Adicionar nova lâmpada")
        self.botaoAddLampada[NomeComodo].configure(command = lambda n=NomeComodo: self.AdicionarLampadaBotao(n))
        self.botaoAddLampada[NomeComodo].pack(side = "bottom",pady= (10,10))
        
        #Ar Condicionado
        self.botaoAr[NomeComodo] = BotaoDispositivo(self.dispositivosFrame[NomeComodo], nomeComodo="Ar Condicionado",
                                  numeroDispositivos= 0,
                                  caminho="src/icons/arCondicionado.png")
        self.botaoAr[NomeComodo].configure(command = lambda n=NomeComodo:self.MudarFrameAr(n))
        self.botaoAr[NomeComodo].pack(side="top", pady= (10,10))
        arCondicionadoFrame = ArCondicionadoFrame(master=self)
        arCondicionadoFrame.header.iconeBotao.configure(command = lambda n=NomeComodo: self.VoltarFrameDispositivosAr(n))
        self.ArCondicionadoFrame[NomeComodo] = arCondicionadoFrame
        
        self.botaoAddAr[NomeComodo] = BotaoAdd(self.ArCondicionadoFrame[NomeComodo], label="Adicionar novo ar")
        self.botaoAddAr[NomeComodo].configure(command = lambda n=NomeComodo: self.AdicionarArBotao(n))
        self.botaoAddAr[NomeComodo].pack(side = "bottom",pady= (10,10))
        
        #Janela
        self.botaoJanela[NomeComodo] = BotaoDispositivo(self.dispositivosFrame[NomeComodo], nomeComodo="Janela",
                                  numeroDispositivos= 0,
                                  caminho="src/icons/Janela.png")
        self.botaoJanela[NomeComodo].configure(command = lambda n=NomeComodo:self.MudarFrameJanela(n))
        self.botaoJanela[NomeComodo].pack(side="top", pady= (10,10))
        
        janelaFrame = JanelaFrame(master=self)
        janelaFrame.header.iconeBotao.configure(command = lambda n=NomeComodo: self.VoltarFrameDispositivosJanela(n))
        self.janelaFrame[NomeComodo] = janelaFrame
        
        self.botaoAddJanela[NomeComodo] = BotaoAdd(self.janelaFrame[NomeComodo], label="Adicionar nova janela")
        self.botaoAddJanela[NomeComodo].configure(command = lambda n=NomeComodo: self.AdicionarJanelaBotao(n))
        self.botaoAddJanela[NomeComodo].pack(side = "bottom",pady= (10,10))
        
        #Cortina
        self.botaoCortina[NomeComodo] = BotaoDispositivo(self.dispositivosFrame[NomeComodo], nomeComodo="Cortina",
                                  numeroDispositivos= 0,
                                  caminho="src/icons/cortina.png")
        self.botaoCortina[NomeComodo].configure(command = lambda n=NomeComodo:self.MudarFrameCortina(n))
        self.botaoCortina[NomeComodo].pack(side="top", pady= (10,10))  
        cortinaFrame = CortinaFrame(master=self)
        cortinaFrame.header.iconeBotao.configure(command = lambda n=NomeComodo: self.VoltarFrameDispositivosCortina(n))
        self.cortinasFrame[NomeComodo] = cortinaFrame 
        
        self.botaoAddCortina[NomeComodo] = BotaoAdd(self.cortinasFrame[NomeComodo], label="Adicionar nova cortina")
        self.botaoAddCortina[NomeComodo].configure(command = lambda n=NomeComodo: self.AdicionarCortinaBotao(n))
        self.botaoAddCortina[NomeComodo].pack(side = "bottom",pady= (10,10))   
        
        self.inputComodoAdd.pack_forget()
        
    def RemoverComodoBotao(self):
        #Função para remover um comodo
        self.inputComodoRemove = ComodoRemoveFrame(self.comodoFrame)
        self.inputComodoRemove.submit.configure(command = lambda:self.SubmeterRemoverComodo())
        self.inputComodoRemove.pack(side="top")
        
    def SubmeterRemoverComodo(self):
        #Verifica se o comodo existe e o remove
        #Se não existir uma mensagem de erro é mostrada
        NomeComodo = self.inputComodoRemove.input.get()
        self.CarregarVetores()
        self.labelInexistente.pack_forget()
        for nomes,numero in self.nomeComodos:
            if nomes == NomeComodo:
                self.casa.RemoverComodo(NomeComodo)
                self.CarregarVetores()
                self.inputComodoRemove.pack_forget()
                self.botoesComodo[NomeComodo].pack_forget()
                self.botoesComodo[NomeComodo].destroy()
                
                #Destruindo o frame para os dispostivos do novo comodo        
                self.dispositivosFrame[NomeComodo].destroy()
                self.lampadasFrame[NomeComodo].destroy()
                self.ArCondicionadoFrame[NomeComodo].destroy()
                self.janelaFrame[NomeComodo].destroy()
                self.cortinasFrame[NomeComodo].destroy()
                return
        self.inputComodoRemove.pack_forget()
        self.labelInexistente.pack(side="top")
        self.inputComodoRemove.pack(side = "top")
        
    def MudarFrameDispositivos(self,nome):
        '''
        Muda a tela atual para a pagina dos dispositivos de um comodo
        :param nome: representa o nome do comodo que se quer acessar
        '''
        self.comodoFrame.pack_forget()
        self.update()
        self.dispositivosFrame[nome].pack(side = "top")
        self.dispositivosFrame[nome].update()
        self.frameAtual = nome
        
    def VoltarFrameComodos(self):
        '''
        Muda a tela atual para a pagina inicial
        '''
        self.dispositivosFrame[self.frameAtual].pack_forget()
        self.update()
        self.comodoFrame.pack(side ='top')
        self.comodoFrame.update()
     
    def MudarFrameLampadas(self,nome):
        '''
        Muda a tela atual para a pagina das lampadas de um comodo
        :param nome: representa o nome do comodo que se quer acessar as lampadas
        '''
        self.dispositivosFrame[self.frameAtual].pack_forget()
        self.update()
        self.lampadasFrame[nome].pack(side ='top')
        self.lampadasFrame[nome].update()
        self.frameAtual = nome
        self.update()
        
    def VoltarFrameDispositivosLampadas(self,nome):
        '''
        Muda a tela atual para a pagina dos dispositivos de um comodo
        :param nome: representa o nome do comodo que se quer acessar os dispositivos
        '''
        self.lampadasFrame[self.frameAtual].pack_forget()
        self.update()
        self.dispositivosFrame[nome].pack(side='top')
        self.dispositivosFrame[nome].update()
        self.frameAtual = nome
        self.update()
        
    def MudarFrameConfigLampadas(self,nome):
        '''
        Muda a tela atual para a pagina de uma lampada especifica de um comodo
        :param nome: representa o nome da lampada que se quer acessar
        '''
        self.lampadasFrame[self.frameAtual].pack_forget()
        self.update()
        self.lampadasConfig[nome].pack(side ='top')
        self.lampadasConfig[nome].update()
        self.frameAtual = nome
        self.update()
        
    def VoltarFrameLampadas(self,nome):
        '''
        Muda a tela atual para a pagina das lampadas de um comodo
        :param nome: representa o nome do comodo que se quer acessar as lampadas
        '''
        self.lampadasConfig[self.frameAtual].pack_forget()
        self.update()
        self.lampadasFrame[nome].pack(side='top')
        self.lampadasFrame[nome].update()
        self.update()
        self.frameAtual = nome
        self.RecarregarBotoesLampada(nome)
        self.CarregarVetores()

        
    def MudarFrameAr(self,nome):
        '''
        Muda a tela atual para a pagina dos Ares Condicionados de um comodo
        :param nome: representa o nome do comodo que se quer acessar os ares condicionados
        '''
        self.dispositivosFrame[self.frameAtual].pack_forget()
        self.update()
        self.ArCondicionadoFrame[nome].pack(side ='top')
        self.ArCondicionadoFrame[nome].update()
        self.frameAtual = nome
        self.update()
        
    def VoltarFrameDispositivosAr(self,nome):
        '''
        Muda a tela atual para a pagina dos dispositivos de um comodo
        :param nome: representa o nome do comodo que se quer acessar
        '''
        self.ArCondicionadoFrame[self.frameAtual].pack_forget()
        self.update()
        self.dispositivosFrame[nome].pack(side='top')
        self.dispositivosFrame[nome].update()
        self.frameAtual = nome
        self.update()
        
    def MudarFrameConfigAr(self,nome):
        '''
        Muda a tela atual para a pagina de um ar condicionado específico
        :param nome: representa o nome do ar condicionado que se quer acessar
        '''
        self.ArCondicionadoFrame[self.frameAtual].pack_forget()
        self.update()
        self.arConfig[nome].pack(side ='top')
        self.arConfig[nome].update()
        self.frameAtual = nome
        self.update()
        
    def VoltarFrameAr(self,nome):
        '''
        Muda a tela atual para a pagina dos ares condicionados de um comodo
        :param nome: representa o nome do comodo que se quer acessar os ares condicionados
        '''
        self.arConfig[self.frameAtual].pack_forget()
        self.update()
        self.ArCondicionadoFrame[nome].pack(side='top')
        self.ArCondicionadoFrame[nome].update()
        self.update()
        self.frameAtual = nome  
        self.RecarregarBotoesAr(nome)
        self.CarregarVetores()
         
    def MudarFrameJanela(self,nome):
        '''
        Muda a tela atual para a pagina das Janelas de um comodo
        :param nome: representa o nome do comodo que se quer acessar as janelas
        '''
        self.dispositivosFrame[self.frameAtual].pack_forget()
        self.update()
        self.janelaFrame[nome].pack(side ='top')
        self.janelaFrame[nome].update()
        self.frameAtual = nome
        self.update()
        
    def VoltarFrameDispositivosJanela(self,nome):
        '''
        Muda a tela atual para a pagina dos dispositivos de um comodo
        :param nome: representa o nome do comodo que se quer acessar
        '''
        self.janelaFrame[self.frameAtual].pack_forget()
        self.update()
        self.dispositivosFrame[nome].pack(side='top')
        self.dispositivosFrame[nome].update()
        self.update()
        self.frameAtual = nome   
    
    def MudarFrameConfigJanela(self,nome):
        '''
        Muda a tela atual para a pagina de uma janela especifica
        :param nome: representa o nome da janela que se quer acessar
        '''
        self.janelaFrame[self.frameAtual].pack_forget()
        self.update()
        self.janelaConfig[nome].pack(side ='top')
        self.janelaConfig[nome].update()
        self.frameAtual = nome
        self.update()
        
    def VoltarFrameJanela(self,nome):
        '''
        Muda a tela atual para a pagina das janelas de um comodo
        :param nome: representa o nome do comodo que se quer acessar as janelas
        '''
        self.janelaConfig[self.frameAtual].pack_forget()
        self.update()
        self.janelaFrame[nome].pack(side='top')
        self.janelaFrame[nome].update()
        self.update()
        self.frameAtual = nome
        self.RecarregarBotoesJanela(nome)
        self.CarregarVetores()
    
    def MudarFrameCortina(self,nome):
        '''
        Muda a tela atual para a pagina das cortinas de um comodo
        :param nome: representa o nome do comodo que se quer acessar a cortinas
        '''
        self.dispositivosFrame[self.frameAtual].pack_forget()
        self.update()
        self.cortinasFrame[nome].pack(side ='top')
        self.cortinasFrame[nome].update()
        self.frameAtual = nome
        self.update()
        
    def VoltarFrameDispositivosCortina(self,nome):
        '''
        Muda a tela atual para a pagina dos dispositivos de um comodo
        :param nome: representa o nome do comodo que se quer acessar
        '''
        self.cortinasFrame[self.frameAtual].pack_forget()
        self.update()
        self.dispositivosFrame[nome].pack(side='top')
        self.dispositivosFrame[nome].update()
        self.update()
        self.frameAtual = nome      
        
    def MudarFrameConfigCortina(self,nome):
        '''
        Muda a tela atual para a pagina de uma cortina 
        :param nome: representa o nome da cortina que se quer acessar
        '''
        self.cortinasFrame[self.frameAtual].pack_forget()
        self.update()
        self.cortinaConfig[nome].pack(side ='top')
        self.cortinaConfig[nome].update()
        self.frameAtual = nome
        self.update()
        
    def VoltarFrameCortina(self,nome):
        '''
        Muda a tela atual para a pagina das cortinas de um comodo
        :param nome: representa o nome do comodo que se quer acessar as cortinas
        '''
        self.cortinaConfig[self.frameAtual].pack_forget()
        self.update()
        self.cortinasFrame[nome].pack(side='top')
        self.cortinasFrame[nome].update()
        self.update()
        self.frameAtual = nome      
        self.RecarregarBotoesCortina(nome)
        self.CarregarVetores()
        
app = App()
app.mainloop()