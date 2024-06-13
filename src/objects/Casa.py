import sys
import os

# Pega o diretório pai do arquivo
diretorioPai = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# fornece o caminho
sys.path.append(diretorioPai)

from openpyxl import Workbook,load_workbook
from Interfaces.Interfaces_casa import InterfaceCasa
from Comodo import Comodo, criar_comodo

wb_Casa_comodo=load_workbook("Casa_comodos.xlsx")
ws=wb_Casa_comodo["Comodos"]

class Casa(InterfaceCasa):
    def __init__(self, nome: str) -> None:
        self.__nome = nome
        self.comodos = {}

    #Retorna o Nome da Casa
    def Nome(self) -> str:
        return self.__nome
    
    #Seta o Nome da Casa 
    def SetNome(self, nomenovo: str) -> None:
        self.__nome = nomenovo

    #Adiciona um Comodo sem nenhum dispositivo nele 
    def AdicionarComodo(self, nomedocomodo: str) -> None:
        if not self.VerificarDuplicado(nomedocomodo):
            ws.append([nomedocomodo,0])
            wb_Casa_comodo.save("Casa_comodos.xlsx")# Salva o workbook após adicionar o cômodo  
        self.comodos[nomedocomodo] = criar_comodo(Comodo, nomedocomodo)
            
    #Remove o Comodo e todos os dispositivos contidos nele 
    def RemoverComodo(self, nomecomodo: str) -> None:
        self.comodos[nomecomodo]=criar_comodo(Comodo,nomecomodo)       
        self.comodos[nomecomodo].ApagarTodosdispositivoscomodo()      
        for i, row in enumerate(ws.iter_rows(), start=2):
                if row[0].value == nomecomodo:
                    ws.delete_rows(i, 1)
        if nomecomodo in self.comodos:
                self.comodos.pop(nomecomodo)
                     
        wb_Casa_comodo.save("Casa_comodos.xlsx") # Salva o workbook após remover o cômodo

    #Verifica se há comodos duplicados
    def VerificarDuplicado(self, nomedocomodo: str) -> bool:
        for row in ws.iter_rows(values_only=True):
            if nomedocomodo in row:
                return True
        return False

    #Salva a quantidade de dispositivos presente nos Comodos
    def SalvarQuantidadeDeDispositivosComodo(self) -> None:
      for row in range(2,(ws.max_row+1)):
        ws['B'+str(row)]=self.comodos[ws['A' + str(row)].value].QuantidadeDispositivo()
        wb_Casa_comodo.save("Casa_comodos.xlsx")
